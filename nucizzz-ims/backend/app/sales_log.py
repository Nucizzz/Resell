from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook

BASE_DIR = Path(__file__).resolve().parent
SALES_DIR = BASE_DIR / "data"
SALES_FILE = SALES_DIR / "vendite.xlsx"

HEADERS = [
    "Timestamp",
    "Location",
    "Product ID",
    "Barcode",
    "Nome",
    "Taglia",
    "Prezzo vendita",
    "Quantità",
]


def _ensure_workbook() -> None:
    SALES_DIR.mkdir(parents=True, exist_ok=True)
    if SALES_FILE.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendite"
    ws.append(HEADERS)
    wb.save(SALES_FILE)


def log_sale(
    *,
    product_id: int,
    barcode: Optional[str],
    title: str,
    size: Optional[str],
    sale_price: Optional[float],
    quantity: int,
    location_name: Optional[str],
) -> None:
    """Append a sale row to the Excel workbook."""
    _ensure_workbook()
    wb = load_workbook(SALES_FILE)
    ws = wb.active
    ws.append(
        [
            datetime.utcnow().isoformat(timespec="seconds"),
            location_name or "",
            product_id,
            barcode or "",
            title,
            size or "",
            sale_price if sale_price is not None else "",
            quantity,
        ]
    )
    wb.save(SALES_FILE)
    wb.close()
